import json
import os
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

API_URL    = "http://localhost:8000/api/test/agent"
EXCEL_FILE = r"F:\Downloads\archive\Test Sample Questions.xlsx"
SRC_SHEET  = "MongoDB Builder"
DST_SHEET  = "Result Summarizer"
BATCH_SIZE = 5

# ── Load rows from the builder sheet ─────────────────────────────────────────
try:
    src = pd.read_excel(EXCEL_FILE, sheet_name=SRC_SHEET)
except (FileNotFoundError, ValueError):
    src = pd.DataFrame(columns=["User Question", "AI Pipeline"])

validRows = src[src["AI Pipeline"].notna()].copy()
validRows = validRows[["User Question", "AI Pipeline"]].reset_index(drop=True)
validRows.index += 1

# ── Build working dataframe ───────────────────────────────────────────────────
df = validRows.copy()
for col in ["AI Answer", "Passed"]:
    if col not in df.columns:
        df[col] = pd.Series(dtype=object)
    else:
        df[col] = df[col].astype(object)

rows = list(df.iterrows())


def callAgent(item):
    i, row = item
    question    = row["User Question"]
    pipelineRaw = row["AI Pipeline"]

    try:
        pipeline = json.loads(pipelineRaw)
    except Exception as e:
        return i, None, f"Pipeline parse error: {e}"

    payload = {
        "agent": "result_summarizer",
        "question": question,
        "pipeline": pipeline,
        # results omitted — API will run the pipeline automatically
        "history": [],
    }
    try:
        response = requests.post(API_URL, json=payload, timeout=120)
        if not response.ok:
            return i, None, f"HTTP {response.status_code}: {response.text[:200]}"
        result = response.json()
        return i, result.get("answer"), None
    except requests.exceptions.JSONDecodeError as e:
        return i, None, f"JSON error: {e} | body: {response.text[:200]}"
    except Exception as e:
        return i, None, str(e)


# ── Run in batches ────────────────────────────────────────────────────────────
for batchStart in range(0, len(rows), BATCH_SIZE):
    batch = rows[batchStart : batchStart + BATCH_SIZE]

    with ThreadPoolExecutor(max_workers=BATCH_SIZE) as executor:
        futures = {executor.submit(callAgent, item): item for item in batch}

        for future in as_completed(futures):
            i, answer, error = future.result()
            if error:
                print(f"✗ {i}: Failed — {error}")
            else:
                df.at[i, "AI Answer"] = answer
                print(f"✓ {i}: {df.at[i, 'User Question']}")

    print(f"  — batch {batchStart // BATCH_SIZE + 1} done ({min(batchStart + BATCH_SIZE, len(rows))}/{len(rows)})")


# ── Reorder: Passed last ──────────────────────────────────────────────────────
df = df[[c for c in df.columns if c != "Passed"] + ["Passed"]]
df.index.name = "#"
df = df.reset_index()

# ── Write to Excel ────────────────────────────────────────────────────────────
writeMode = "a" if os.path.exists(EXCEL_FILE) else "w"
with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode=writeMode, if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name=DST_SHEET, index=False)

    ws = writer.sheets[DST_SHEET]

    # ── Styles ────────────────────────────────────────────────────────────
    baseHeaderFill   = PatternFill("solid", fgColor="2D2D2D")
    aiHeaderFill     = PatternFill("solid", fgColor="145A32")
    passedHeaderFill = PatternFill("solid", fgColor="6E2F8E")
    aiFill           = PatternFill("solid", fgColor="EAF4FB")
    passedGreenFill  = PatternFill("solid", fgColor="B7E4C7")
    passedRedFill    = PatternFill("solid", fgColor="F5B7B1")
    passedEmptyFill  = PatternFill("solid", fgColor="FEF9E7")
    headerFont       = Font(bold=True, color="FFFFFF", size=10)
    bodyFont         = Font(size=10)
    thinSide         = Side(style="thin", color="CCCCCC")
    thinBorder       = Border(left=thinSide, right=thinSide, top=thinSide, bottom=thinSide)
    centerAlign      = Alignment(horizontal="center", vertical="top", wrap_text=False)
    leftAlign        = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    AI_COLS = {"AI Answer"}

    headers        = [cell.value for cell in ws[1]]
    aiColIndices   = {col for col, name in enumerate(headers, 1) if name in AI_COLS}
    passedColIndex = next((col for col, name in enumerate(headers, 1) if name == "Passed"), None)

    # ── Header row ────────────────────────────────────────────────────────
    for cell in ws[1]:
        name = cell.value
        if name in AI_COLS:
            cell.fill = aiHeaderFill
        elif name == "Passed":
            cell.fill = passedHeaderFill
        else:
            cell.fill = baseHeaderFill
        cell.font      = headerFont
        cell.border    = thinBorder
        cell.alignment = centerAlign

    # ── Body rows ─────────────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font   = bodyFont
            cell.border = thinBorder

            isAiCol     = cell.column in aiColIndices
            isPassedCol = cell.column == passedColIndex

            if isPassedCol and cell.value is True:
                cell.fill      = passedGreenFill
                cell.alignment = centerAlign
            elif isPassedCol and cell.value is False:
                cell.fill      = passedRedFill
                cell.alignment = centerAlign
            elif isPassedCol:
                cell.fill      = passedEmptyFill
                cell.alignment = centerAlign
            elif isAiCol:
                cell.fill      = aiFill
                cell.alignment = leftAlign
            else:
                cell.alignment = leftAlign

    ws.auto_filter.ref = ws.dimensions

    if passedColIndex:
        passedLetter = get_column_letter(passedColIndex)
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True, showDropDown=False)
        dv.sqref = f"{passedLetter}2:{passedLetter}{ws.max_row}"
        ws.add_data_validation(dv)

    for col in ws.columns:
        colLetter = get_column_letter(col[0].column)
        maxLen = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[colLetter].width = min(maxLen + 4, 80)

print("Done. Result Summarizer sheet populated.")

"""
Builds (or refreshes) the "Accuracy Summary" sheet in the test Excel file.

For each tracked sheet it reads the header row to find which column is "Passed"
and which column is "User Question", then writes live Excel COUNTIF formulas so
the summary table updates automatically whenever you fill in the Passed cells.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

EXCEL_FILE    = r"F:\Downloads\archive\Test Sample Questions.xlsx"
SUMMARY_SHEET = "Accuracy Summary"

# Ordered list of sheets to track: (sheet_name, display_label)
TRACKED_SHEETS = [
    ("User Query Validator", "User Query Validator"),
    ("MongoDB Builder",      "MongoDB Builder"),
    ("Query Validator",      "Query Validator"),
    ("Result Summarizer",    "Result Summarizer"),
]


def findColumnLetter(ws, columnName: str):
    """Return the column letter for a given header value in row 1, or None."""
    for cell in ws[1]:
        if cell.value == columnName:
            return get_column_letter(cell.column)
    return None


# ── Load or create workbook ───────────────────────────────────────────────────
if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
else:
    wb = Workbook()

# Remove and recreate the summary sheet so it's always fresh
if SUMMARY_SHEET in wb.sheetnames:
    del wb[SUMMARY_SHEET]
ws = wb.create_sheet(SUMMARY_SHEET, 0)   # pin as the first/leftmost tab


# ── Styles ────────────────────────────────────────────────────────────────────
titleFill        = PatternFill("solid", fgColor="2D2D2D")
passedFill       = PatternFill("solid", fgColor="145A32")
failedFill       = PatternFill("solid", fgColor="922B21")
pendingFill      = PatternFill("solid", fgColor="7D6608")
accuracyFill     = PatternFill("solid", fgColor="154360")
rowEvenFill      = PatternFill("solid", fgColor="F2F3F4")
rowOddFill       = PatternFill("solid", fgColor="FFFFFF")
notFoundFill     = PatternFill("solid", fgColor="F9EBEA")
headerFont       = Font(bold=True, color="FFFFFF", size=11)
bodyFont         = Font(size=11)
boldBodyFont     = Font(size=11, bold=True)
thinSide         = Side(style="thin", color="CCCCCC")
thinBorder       = Border(left=thinSide, right=thinSide, top=thinSide, bottom=thinSide)
centerAlign      = Alignment(horizontal="center", vertical="center")
leftAlign        = Alignment(horizontal="left",   vertical="center")
rightAlign       = Alignment(horizontal="right",  vertical="center")


# ── Column definitions ────────────────────────────────────────────────────────
# (header label, fill, alignment, width)
COLUMNS = [
    ("Sheet",      titleFill,    leftAlign,   30),
    ("Total",      titleFill,    centerAlign, 12),
    ("Passed",     passedFill,   centerAlign, 12),
    ("Failed",     failedFill,   centerAlign, 12),
    ("Pending",    pendingFill,  centerAlign, 12),
    ("Accuracy",   accuracyFill, centerAlign, 14),
]

# ── Write header row ──────────────────────────────────────────────────────────
for col, (label, fill, align, width) in enumerate(COLUMNS, start=1):
    cell           = ws.cell(row=1, column=col, value=label)
    cell.fill      = fill
    cell.font      = headerFont
    cell.border    = thinBorder
    cell.alignment = align
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# ── Write one row per tracked sheet ──────────────────────────────────────────
for rowNum, (sheetName, label) in enumerate(TRACKED_SHEETS, start=2):
    rowFill = rowEvenFill if rowNum % 2 == 0 else rowOddFill

    def writeCell(col, value, numFmt=None, font=None, fill=None, align=None):
        c            = ws.cell(row=rowNum, column=col, value=value)
        c.fill       = fill or rowFill
        c.font       = font or bodyFont
        c.border     = thinBorder
        c.alignment  = align or centerAlign
        if numFmt:
            c.number_format = numFmt
        return c

    if sheetName not in wb.sheetnames:
        writeCell(1, label,           align=leftAlign, font=Font(size=11, italic=True, color="999999"))
        writeCell(2, "Sheet not found yet")
        writeCell(3, "—")
        writeCell(4, "—")
        writeCell(5, "—")
        writeCell(6, "—")
        continue

    src          = wb[sheetName]
    passedLetter = findColumnLetter(src, "Passed")
    questionLetter = findColumnLetter(src, "User Question")

    if not passedLetter or not questionLetter:
        writeCell(1, label, align=leftAlign)
        writeCell(2, "Column not found")
        writeCell(3, "—"); writeCell(4, "—"); writeCell(5, "—"); writeCell(6, "—")
        continue

    maxRow = 10000  # safe upper bound for formulas
    pRange = f"'{sheetName}'!${passedLetter}$2:${passedLetter}${maxRow}"
    qRange = f"'{sheetName}'!${questionLetter}$2:${questionLetter}${maxRow}"

    # Row reference letters for cross-formula use
    bLetter = get_column_letter(2)  # Total  → col B
    cLetter = get_column_letter(3)  # Passed → col C
    dLetter = get_column_letter(4)  # Failed → col D

    writeCell(1, label, align=leftAlign, font=boldBodyFont)
    writeCell(2, f"=COUNTA({qRange})")                                         # Total
    writeCell(3, f"=COUNTIF({pRange},TRUE)")                                   # Passed
    writeCell(4, f"=COUNTIF({pRange},FALSE)")                                  # Failed
    writeCell(5, f"={bLetter}{rowNum}-{cLetter}{rowNum}-{dLetter}{rowNum}")    # Pending
    writeCell(                                                                  # Accuracy %
        6,
        f'=IF({cLetter}{rowNum}+{dLetter}{rowNum}>0,'
        f'{cLetter}{rowNum}/({cLetter}{rowNum}+{dLetter}{rowNum}),"")',
        numFmt=FORMAT_PERCENTAGE_00,
    )

ws.row_dimensions[rowNum].height = 20

# ── Totals row ────────────────────────────────────────────────────────────────
totalRow    = len(TRACKED_SHEETS) + 2
totalsFill  = PatternFill("solid", fgColor="1A252F")

def writeTotalCell(col, value, numFmt=None):
    c               = ws.cell(row=totalRow, column=col, value=value)
    c.fill          = totalsFill
    c.font          = Font(bold=True, color="FFFFFF", size=11)
    c.border        = thinBorder
    c.alignment     = centerAlign
    if numFmt:
        c.number_format = numFmt
    return c

dataStart = 2
dataEnd   = totalRow - 1
writeTotalCell(1, "TOTAL", )
writeTotalCell(2, f"=SUM(B{dataStart}:B{dataEnd})")
writeTotalCell(3, f"=SUM(C{dataStart}:C{dataEnd})")
writeTotalCell(4, f"=SUM(D{dataStart}:D{dataEnd})")
writeTotalCell(5, f"=SUM(E{dataStart}:E{dataEnd})")
writeTotalCell(
    6,
    f'=IF(C{totalRow}+D{totalRow}>0,C{totalRow}/(C{totalRow}+D{totalRow}),"")',
    numFmt=FORMAT_PERCENTAGE_00,
)
ws.row_dimensions[totalRow].height = 22

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(EXCEL_FILE)
print("Done. Accuracy Summary sheet updated.")

import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import os

API_URL = "http://localhost:8000/api/test/agent"
EXCEL_FILE = r"F:\Downloads\archive\Test Sample Questions.xlsx"
SHEET_NAME = "User Query Validator"
BATCH_SIZE = 5

try:
    df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
except (FileNotFoundError, ValueError):
    df = pd.DataFrame(columns=["User Question"])

# Ensure AI columns accept any type (bool, str, None)
for col in ["AI isValid", "AI Normalized Query", "AI Clarifying Question", "Passed"]:
    df[col] = df[col].astype(object) if col in df.columns else pd.Series(dtype=object)

rows = list(df.iterrows())


def callAgent(item):
    i, row = item
    question = row["User Question"]
    payload = {
        "agent": "user_query_validator",
        "message": question,
        "history": [],
    }
    try:
        response = requests.post(API_URL, json=payload, timeout=60)
        if not response.ok:
            return i, question, None, None, None, f"HTTP {response.status_code}: {response.text[:200]}"
        result = response.json()
        return i, question, result.get("isValid"), result.get("normalizedQuery"), result.get("clarifyingQuestion"), None
    except requests.exceptions.JSONDecodeError as e:
        return i, question, None, None, None, f"JSON error: {e} | body: {response.text[:200]}"
    except Exception as e:
        return i, question, None, None, None, str(e)


# Process in batches of BATCH_SIZE
for batchStart in range(0, len(rows), BATCH_SIZE):
    batch = rows[batchStart : batchStart + BATCH_SIZE]

    with ThreadPoolExecutor(max_workers=BATCH_SIZE) as executor:
        futures = {executor.submit(callAgent, item): item for item in batch}

        for future in as_completed(futures):
            i, question, isValid, normalizedQuery, clarifyingQuestion, error = future.result()
            if error:
                print(f"✗ {i+1}: Failed → {error}")
            else:
                df.at[i, "AI isValid"] = isValid
                df.at[i, "AI Normalized Query"] = normalizedQuery
                df.at[i, "AI Clarifying Question"] = clarifyingQuestion
                print(f"✓ {i+1}: {question}")

    print(f"  — batch {batchStart // BATCH_SIZE + 1} done ({min(batchStart + BATCH_SIZE, len(rows))}/{len(rows)})")

# Leave Passed empty for manual review

# Ensure "Passed" is the last column
cols = [c for c in df.columns if c != "Passed"] + ["Passed"]
df = df[cols]

# Save back to same file
writeMode = "a" if os.path.exists(EXCEL_FILE) else "w"
with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode=writeMode, if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    ws = writer.sheets[SHEET_NAME]

    # ── Styles ────────────────────────────────────────────────────────────
    headerFill         = PatternFill("solid", fgColor="2D2D2D")   # base columns
    expectedHeaderFill = PatternFill("solid", fgColor="1A5276")   # expected group  – deep blue
    aiHeaderFill       = PatternFill("solid", fgColor="145A32")   # AI results group – deep green
    passedHeaderFill   = PatternFill("solid", fgColor="6E2F8E")   # passed column    – deep purple
    aiFill             = PatternFill("solid", fgColor="EAF4FB")   # light blue for AI columns
    greenFill          = PatternFill("solid", fgColor="D6F5D6")   # True
    redFill            = PatternFill("solid", fgColor="FAD7D7")   # False
    headerFont         = Font(bold=True, color="FFFFFF", size=10)
    bodyFont           = Font(size=10)
    thinSide           = Side(style="thin", color="CCCCCC")
    thinBorder         = Border(left=thinSide, right=thinSide, top=thinSide, bottom=thinSide)
    centerAlign        = Alignment(horizontal="center", vertical="top",    wrap_text=False)
    leftAlign          = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    EXPECTED_COLS = {"Should Be Valid", "Expected Normalized Query", "Expected Clarifying Question"}
    AI_COLS       = {"AI isValid", "AI Normalized Query", "AI Clarifying Question"}

    headers        = [cell.value for cell in ws[1]]
    aiColIndices   = {col for col, name in enumerate(headers, 1) if isinstance(name, str) and name.startswith("AI ")}
    passedColIndex = next((col for col, name in enumerate(headers, 1) if name == "Passed"), None)
    passedGreenFill = PatternFill("solid", fgColor="B7E4C7")  # distinct green for Passed=True
    passedRedFill   = PatternFill("solid", fgColor="F5B7B1")  # distinct red  for Passed=False
    passedEmptyFill = PatternFill("solid", fgColor="FEF9E7")  # soft yellow   for Passed=empty

    # ── Header row ────────────────────────────────────────────────────────
    for cell in ws[1]:
        name = cell.value
        if name in EXPECTED_COLS:
            cell.fill = expectedHeaderFill
        elif name in AI_COLS:
            cell.fill = aiHeaderFill
        elif name == "Passed":
            cell.fill = passedHeaderFill
        else:
            cell.fill = headerFill
        cell.font      = headerFont
        cell.border    = thinBorder
        cell.alignment = centerAlign

    # ── Body rows ─────────────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font   = bodyFont
            cell.border = thinBorder

            isAiCol = cell.column in aiColIndices

            isPassedCol = cell.column == passedColIndex

            if isPassedCol and cell.value is True:
                cell.fill      = passedGreenFill
                cell.alignment = centerAlign
            elif isPassedCol and cell.value is False:
                cell.fill      = passedRedFill
                cell.alignment = centerAlign
            elif isPassedCol:
                cell.fill      = passedEmptyFill
                cell.alignment = centerAlign
            elif cell.value is True:
                cell.fill      = greenFill
                cell.alignment = centerAlign
            elif cell.value is False:
                cell.fill      = redFill
                cell.alignment = centerAlign
            elif isAiCol:
                cell.fill      = aiFill
                cell.alignment = leftAlign
            else:
                cell.alignment = leftAlign

    # ── Auto-filter (column headers act as table filters) ─────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Passed column: True/False dropdown ───────────────────────────────
    if passedColIndex:
        passedLetter = get_column_letter(passedColIndex)
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True, showDropDown=False)
        dv.sqref = f"{passedLetter}2:{passedLetter}{ws.max_row}"
        ws.add_data_validation(dv)

    # ── Auto-fit column widths (generous, capped at 80) ──────────────────
    for col in ws.columns:
        colLetter = get_column_letter(col[0].column)
        maxLen = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[colLetter].width = min(maxLen + 4, 80)

print("Done. AI columns populated.")